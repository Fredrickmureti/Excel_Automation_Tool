#!/bin/bash
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('python-spreadsheets.xlsx')
sheet = wb['python-spreadsheets.xlsx']

for row in range(2, sheet.max_row + 1):
     cell = sheet.cell(row, column_reference)

     values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=x, max_col=y) #in this case x and y represent the min and max colomns respectively

chart = BarChart()
chart.add_data(values)
sheet.add_chart(BarChart, 'e2') 
wb.save('python-spreadsheet2.xlsx')
